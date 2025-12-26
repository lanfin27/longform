"""
엑셀 추출 유틸리티 v2.0

영상 리서치 결과를 Excel/CSV로 내보내기

v2.0 추가:
- 새로운 성과 지표 (전일 기여도, 시간당 조회수, 참여율)
- 스타일링 개선
- 파일명 생성 유틸리티
"""
import pandas as pd
from typing import List, Union
from io import BytesIO
from datetime import datetime


def export_videos_to_excel(
    videos: List[dict],
    filename_prefix: str = "영상_리서치"
) -> BytesIO:
    """
    영상 목록을 엑셀 파일로 변환

    Args:
        videos: VideoInfo.to_excel_row() 형식의 딕셔너리 리스트
        filename_prefix: 파일명 접두어

    Returns:
        BytesIO: 엑셀 파일 바이트 스트림
    """
    if not videos:
        # 빈 파일 반환
        output = BytesIO()
        df = pd.DataFrame()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='영상 목록', index=False)
        output.seek(0)
        return output

    # DataFrame 생성
    df = pd.DataFrame(videos)

    # 컬럼 순서 정렬
    column_order = [
        "영상 제목",
        "영상 URL",
        "영상 유형",
        "영상 길이",
        "업로드일",
        "조회수",
        "좋아요",
        "댓글수",
        "채널명",
        "채널 URL",
        "구독자수",
        "채널 개설일",
        "채널 총 영상수",
        "구독자 대비 조회수",
        "참여율(%)",
        "업로드 후 일수",
        "일일 평균 조회수",
        "급등 점수",
    ]

    # 존재하는 컬럼만 선택
    available_columns = [col for col in column_order if col in df.columns]
    # 나머지 컬럼도 추가
    remaining = [col for col in df.columns if col not in available_columns]
    df = df[available_columns + remaining]

    # 엑셀 파일 생성
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='영상 목록', index=False)

        # 워크시트 스타일링
        worksheet = writer.sheets['영상 목록']

        # 열 너비 자동 조정
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            # openpyxl은 1-indexed, 열 문자로 변환
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output.seek(0)
    return output


def get_column_letter(idx: int) -> str:
    """열 인덱스를 엑셀 열 문자로 변환 (1=A, 27=AA)"""
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def export_to_csv(videos: List[dict]) -> str:
    """CSV 형식으로 변환"""
    if not videos:
        return ""

    df = pd.DataFrame(videos)
    return df.to_csv(index=False, encoding='utf-8-sig')


def export_videos_simple(
    videos: List[dict],
    include_columns: List[str] = None
) -> BytesIO:
    """
    간단한 형식으로 엑셀 내보내기

    Args:
        videos: 영상 데이터 리스트
        include_columns: 포함할 컬럼 목록 (None이면 전체)

    Returns:
        BytesIO: 엑셀 파일 바이트 스트림
    """
    df = pd.DataFrame(videos)

    if include_columns:
        available = [c for c in include_columns if c in df.columns]
        df = df[available]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='영상 목록', index=False)
    output.seek(0)

    return output


def create_summary_sheet(videos: List[dict]) -> pd.DataFrame:
    """요약 통계 시트 생성"""
    if not videos:
        return pd.DataFrame()

    df = pd.DataFrame(videos)

    summary = {
        "항목": [],
        "값": []
    }

    # 전체 통계
    summary["항목"].append("총 영상 수")
    summary["값"].append(len(videos))

    if "조회수" in df.columns:
        summary["항목"].append("총 조회수")
        summary["값"].append(f"{df['조회수'].sum():,}")
        summary["항목"].append("평균 조회수")
        summary["값"].append(f"{df['조회수'].mean():,.0f}")
        summary["항목"].append("최대 조회수")
        summary["값"].append(f"{df['조회수'].max():,}")

    if "구독자수" in df.columns:
        summary["항목"].append("평균 구독자")
        summary["값"].append(f"{df['구독자수'].mean():,.0f}")

    if "영상 유형" in df.columns:
        shorts_count = len(df[df["영상 유형"] == "쇼츠"])
        longform_count = len(df[df["영상 유형"] == "롱폼"])
        summary["항목"].append("롱폼 영상 수")
        summary["값"].append(longform_count)
        summary["항목"].append("쇼츠 영상 수")
        summary["값"].append(shorts_count)

    return pd.DataFrame(summary)


def export_with_summary(
    videos: List[dict],
    filename_prefix: str = "영상_리서치"
) -> BytesIO:
    """
    요약 시트를 포함한 엑셀 파일 생성

    Args:
        videos: 영상 데이터 리스트
        filename_prefix: 파일명 접두어

    Returns:
        BytesIO: 엑셀 파일 바이트 스트림
    """
    if not videos:
        output = BytesIO()
        df = pd.DataFrame()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='영상 목록', index=False)
        output.seek(0)
        return output

    df = pd.DataFrame(videos)
    summary_df = create_summary_sheet(videos)

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 요약 시트
        summary_df.to_excel(writer, sheet_name='요약', index=False)

        # 영상 목록 시트
        df.to_excel(writer, sheet_name='영상 목록', index=False)

        # 열 너비 조정
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(worksheet.iter_cols(min_row=1, max_row=1), 1):
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = 15

    output.seek(0)
    return output


# ==================== v2.0 추가 함수 ====================

def get_excel_filename(keyword: str = "results") -> str:
    """Excel 파일명 생성"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    # 안전한 파일명으로 변환
    import re
    safe_keyword = re.sub(r'[^a-zA-Z0-9가-힣]', '_', keyword)[:20]
    return f"youtube_{safe_keyword}_{timestamp}.xlsx"


def export_with_metrics(
    videos: List[dict],
    keyword: str = "검색결과"
) -> BytesIO:
    """
    성과 지표가 포함된 Excel 내보내기

    Args:
        videos: YouTubeService 형식의 영상 데이터 리스트
        keyword: 검색 키워드 (시트명에 사용)

    Returns:
        BytesIO: 엑셀 파일 바이트 스트림
    """
    if not videos:
        output = BytesIO()
        df = pd.DataFrame()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='검색결과', index=False)
        output.seek(0)
        return output

    # 컬럼 매핑 (YouTubeService 형식 -> 한글)
    column_mapping = {
        'title': '영상 제목',
        'channel_title': '채널명',
        'view_count': '조회수',
        'like_count': '좋아요',
        'comment_count': '댓글수',
        'subscriber_count': '구독자수',
        'duration_formatted': '영상 길이',
        'published_at': '업로드일',
        'video_type': '영상 유형',
        'views_per_hour': '시간당 조회수',
        'daily_contribution': '전일 기여도(%)',
        'engagement_rate': '참여율(%)',
        'views_per_subscriber': '구독자 대비 조회수',
        'days_since_upload': '업로드 후 일수',
        'daily_views': '일일 평균 조회수',
        'video_url': '영상 URL',
        'channel_url': '채널 URL',
    }

    # DataFrame 생성
    df = pd.DataFrame(videos)

    # 영상 유형 한글화
    if 'video_type' in df.columns:
        df['video_type'] = df['video_type'].apply(
            lambda x: '쇼츠' if x == 'shorts' else '롱폼'
        )

    # 업로드일 포맷팅
    if 'published_at' in df.columns:
        df['published_at'] = df['published_at'].apply(
            lambda x: x[:10] if isinstance(x, str) and len(x) >= 10 else x
        )

    # 존재하는 컬럼만 선택 및 이름 변경
    available_columns = [col for col in column_mapping.keys() if col in df.columns]
    df_export = df[available_columns].copy()
    df_export = df_export.rename(columns=column_mapping)

    # 순번 추가
    df_export.insert(0, '순번', range(1, len(df_export) + 1))

    # 요약 통계 생성
    summary_data = {
        '항목': [
            '총 영상 수',
            '총 조회수',
            '평균 조회수',
            '평균 좋아요',
            '평균 참여율',
            '평균 전일 기여도',
            '롱폼 영상 수',
            '쇼츠 영상 수',
        ],
        '값': [
            len(videos),
            f"{sum(v.get('view_count', 0) for v in videos):,}",
            f"{sum(v.get('view_count', 0) for v in videos) // max(len(videos), 1):,}",
            f"{sum(v.get('like_count', 0) for v in videos) // max(len(videos), 1):,}",
            f"{sum(v.get('engagement_rate', 0) for v in videos) / max(len(videos), 1):.2f}%",
            f"{sum(v.get('daily_contribution', 0) for v in videos) / max(len(videos), 1):.2f}%",
            len([v for v in videos if v.get('video_type') != 'shorts']),
            len([v for v in videos if v.get('video_type') == 'shorts']),
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    # Excel 생성
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 요약 시트
        summary_df.to_excel(writer, sheet_name='요약', index=False)

        # 영상 목록 시트
        df_export.to_excel(writer, sheet_name='검색결과', index=False)

        # 스타일링
        _apply_enhanced_styles(writer, '요약', summary_df)
        _apply_enhanced_styles(writer, '검색결과', df_export)

    output.seek(0)
    return output


def _apply_enhanced_styles(writer, sheet_name: str, df: pd.DataFrame):
    """향상된 Excel 스타일 적용"""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter as openpyxl_get_col

        worksheet = writer.sheets[sheet_name]

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 헤더 행 스타일링
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 데이터 행 스타일링
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        # 열 너비 자동 조정
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).apply(len).max() if len(df) > 0 else 0,
                len(str(col))
            )
            col_letter = openpyxl_get_col(idx)
            worksheet.column_dimensions[col_letter].width = min(max_length + 3, 50)

        # 특정 열 너비 조정
        for idx, col_name in enumerate(df.columns, 1):
            col_letter = openpyxl_get_col(idx)
            if '제목' in col_name:
                worksheet.column_dimensions[col_letter].width = 50
            elif 'URL' in col_name:
                worksheet.column_dimensions[col_letter].width = 40

    except ImportError:
        pass  # openpyxl 스타일링 미지원
