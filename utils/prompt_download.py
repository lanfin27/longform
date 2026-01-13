# -*- coding: utf-8 -*-
"""
프롬프트 선택적 다운로드 유틸리티 (v1.0)

씬별 프롬프트를 선택적으로 다운로드하는 기능
- 범위 선택, 개별 선택, 이미지 있는 씬만 선택 지원
- 최신 이미지 메타데이터에서 프롬프트 추출 지원
"""

import json
import re
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Any, Set
from io import BytesIO
from datetime import datetime


# ============================================================
# 최신 이미지 검색 함수
# ============================================================

def get_latest_image_per_scene(backgrounds_folder: Path) -> Dict[int, Dict[str, Any]]:
    """각 씬별 가장 최신 이미지 반환

    파일명 패턴: bg_scene_NNN_TIMESTAMP.png

    Args:
        backgrounds_folder: 이미지 폴더 경로

    Returns:
        {
            scene_num: {
                "path": str,
                "timestamp": int,
                "metadata": dict or None
            }
        }
    """

    latest_images = {}
    backgrounds_folder = Path(backgrounds_folder)

    if not backgrounds_folder.exists():
        print(f"[PromptDownload] 폴더 없음: {backgrounds_folder}", flush=True)
        return latest_images

    # 모든 이미지 파일 스캔
    for img_path in backgrounds_folder.glob("bg_scene_*.png"):
        # 파일명에서 씬 번호와 타임스탬프 추출
        # bg_scene_005_1768001279723.png
        match = re.match(r'bg_scene_(\d+)_(\d+)\.png', img_path.name)
        if not match:
            continue

        scene_num = int(match.group(1))
        timestamp = int(match.group(2))

        # 해당 씬의 기존 최신보다 새로우면 교체
        if scene_num not in latest_images or timestamp > latest_images[scene_num]['timestamp']:
            # 메타데이터 로드
            json_path = img_path.with_suffix('.json')
            metadata = None
            if json_path.exists():
                try:
                    with open(json_path, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                except Exception as e:
                    print(f"[PromptDownload] 메타데이터 로드 실패: {json_path} - {e}", flush=True)

            latest_images[scene_num] = {
                'path': str(img_path),
                'timestamp': timestamp,
                'metadata': metadata,
                'metadata_path': str(json_path) if metadata else None
            }

    print(f"[PromptDownload] 최신 이미지: {len(latest_images)}개 씬", flush=True)
    return latest_images


def get_scenes_with_images(backgrounds_folder: Path) -> Set[int]:
    """이미지가 있는 씬 번호 집합 반환"""
    latest = get_latest_image_per_scene(backgrounds_folder)
    return set(latest.keys())


# ============================================================
# 프롬프트 추출 함수
# ============================================================

def get_prompt_for_scene(
    scene_num: int,
    scene_data: dict,
    latest_images: dict,
    use_latest_metadata: bool = True,
    style_prefix: str = "",
    style_suffix: str = ""
) -> Dict[str, Any]:
    """씬의 프롬프트 추출 - 최신 이미지 메타데이터 우선

    Args:
        scene_num: 씬 번호
        scene_data: 씬 분석 데이터
        latest_images: get_latest_image_per_scene()의 결과
        use_latest_metadata: True면 메타데이터 우선, False면 씬 분석 데이터만 사용
        style_prefix: 스타일 prefix
        style_suffix: 스타일 suffix

    Returns:
        {
            "scene_number": int,
            "narration": str,
            "original_prompt": str,  # 씬 분석 원본
            "korean_prompt": str,    # 한글 프롬프트
            "final_prompt": str,     # 최종 사용된 프롬프트 (메타데이터에서)
            "negative_prompt": str,
            "source": "metadata" | "scene_analysis",
            "has_image": bool,
            "image_timestamp": int or None
        }
    """

    # 기본값: 씬 분석 데이터에서 추출
    original = scene_data.get("image_prompt_en", "") or scene_data.get("prompts", {}).get("image_prompt_en", "")
    korean = scene_data.get("image_prompt_korean_text", "") or scene_data.get("prompts", {}).get("image_prompt_korean_text", "")
    narration = scene_data.get("narration", scene_data.get("script_text", ""))

    result = {
        "scene_number": scene_num,
        "narration": narration,
        "original_prompt": original,
        "korean_prompt": korean,
        "final_prompt": "",
        "negative_prompt": "",
        "source": "scene_analysis",
        "has_image": False,
        "image_timestamp": None
    }

    # 최신 이미지가 있는지 확인
    if scene_num in latest_images:
        result["has_image"] = True
        result["image_timestamp"] = latest_images[scene_num].get('timestamp')

        # 메타데이터 사용 옵션이 켜져 있으면
        if use_latest_metadata:
            metadata = latest_images[scene_num].get('metadata')

            if metadata and 'prompts' in metadata:
                prompt_data = metadata['prompts']

                # 메타데이터에서 프롬프트 추출
                if prompt_data.get("final"):
                    result["final_prompt"] = prompt_data.get("final", "")
                    result["source"] = "metadata"

                if prompt_data.get("original") and not result["original_prompt"]:
                    result["original_prompt"] = prompt_data.get("original", "")

                if prompt_data.get("negative"):
                    result["negative_prompt"] = prompt_data.get("negative", "")

    # final_prompt가 없으면 원본 + 스타일로 생성
    if not result["final_prompt"] and result["original_prompt"]:
        if style_prefix or style_suffix:
            parts = [style_prefix, result["original_prompt"], style_suffix]
            result["final_prompt"] = ", ".join(p for p in parts if p)
        else:
            result["final_prompt"] = result["original_prompt"]

    return result


def collect_prompts_for_selected_scenes(
    selected_scene_nums: List[int],
    scenes: List[dict],
    latest_images: Dict[int, dict],
    use_latest_metadata: bool = True,
    style_prefix: str = "",
    style_suffix: str = ""
) -> List[Dict[str, Any]]:
    """선택된 씬들의 프롬프트 수집

    Args:
        selected_scene_nums: 선택된 씬 번호 리스트 [1, 2, 5, 10, ...]
        scenes: 전체 씬 데이터
        latest_images: 씬별 최신 이미지 정보
        use_latest_metadata: 최신 메타데이터 사용 여부
        style_prefix: 스타일 prefix
        style_suffix: 스타일 suffix

    Returns:
        프롬프트 데이터 리스트
    """

    prompts = []

    for scene_num in sorted(selected_scene_nums):
        # 씬 데이터 찾기 (scene_id 기반 또는 인덱스 기반)
        scene_data = None

        # scene_id로 찾기
        for s in scenes:
            s_id = s.get("scene_id") or s.get("scene_num") or s.get("id")
            if s_id == scene_num:
                scene_data = s
                break

        # 못 찾으면 인덱스로 (1-based)
        if scene_data is None:
            scene_idx = scene_num - 1
            if 0 <= scene_idx < len(scenes):
                scene_data = scenes[scene_idx]

        if scene_data:
            prompt_data = get_prompt_for_scene(
                scene_num=scene_num,
                scene_data=scene_data,
                latest_images=latest_images,
                use_latest_metadata=use_latest_metadata,
                style_prefix=style_prefix,
                style_suffix=style_suffix
            )
            prompts.append(prompt_data)

    return prompts


def get_prompt_stats(prompts: List[Dict[str, Any]]) -> Dict[str, int]:
    """프롬프트 통계 계산"""

    return {
        "total": len(prompts),
        "with_original": sum(1 for p in prompts if p.get("original_prompt")),
        "with_korean": sum(1 for p in prompts if p.get("korean_prompt")),
        "with_final": sum(1 for p in prompts if p.get("final_prompt")),
        "with_image": sum(1 for p in prompts if p.get("has_image")),
        "from_metadata": sum(1 for p in prompts if p.get("source") == "metadata"),
        "from_analysis": sum(1 for p in prompts if p.get("source") == "scene_analysis")
    }


# ============================================================
# 다운로드 파일 생성 함수
# ============================================================

def generate_prompts_excel(
    prompts: List[Dict[str, Any]],
    include_source: bool = True,
    include_negative: bool = False
) -> bytes:
    """선택된 씬들의 프롬프트 엑셀 생성

    Args:
        prompts: collect_prompts_for_selected_scenes()의 반환값
        include_source: 프롬프트 출처 컬럼 포함 여부
        include_negative: 네거티브 프롬프트 포함 여부

    Returns:
        엑셀 파일 바이트
    """

    import pandas as pd

    data = []

    for p in prompts:
        row = {
            "씬번호": p["scene_number"],
            "나레이션": (p.get("narration", "")[:100] + "...") if len(p.get("narration", "")) > 100 else p.get("narration", ""),
            "원본프롬프트": p.get("original_prompt", ""),
            "한글프롬프트": p.get("korean_prompt", ""),
            "최종프롬프트": p.get("final_prompt", ""),
        }

        if include_negative:
            row["네거티브"] = p.get("negative_prompt", "")

        if include_source:
            source_text = "메타데이터" if p.get("source") == "metadata" else "씬분석"
            row["출처"] = source_text
            row["이미지유무"] = "O" if p.get("has_image") else "X"

        data.append(row)

    df = pd.DataFrame(data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='프롬프트')

    output.seek(0)
    return output.getvalue()


def generate_prompts_zip(
    prompts: List[Dict[str, Any]],
    prompt_type: str = "final"  # "final" | "korean" | "original"
) -> bytes:
    """선택된 씬들의 프롬프트 ZIP 생성

    Args:
        prompts: 프롬프트 데이터 리스트
        prompt_type: 다운로드할 프롬프트 종류

    Returns:
        ZIP 파일 바이트
    """

    zip_buffer = BytesIO()
    file_count = 0

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in prompts:
            scene_num = p["scene_number"]

            if prompt_type == "korean":
                content = p.get("korean_prompt", "")
                suffix = "korean"
            elif prompt_type == "original":
                content = p.get("original_prompt", "")
                suffix = "original"
            else:  # final
                content = p.get("final_prompt", "")
                suffix = "final"

            if content:
                filename = f"scene_{scene_num:03d}_{suffix}.txt"
                zf.writestr(filename, content)
                file_count += 1

    zip_buffer.seek(0)
    return zip_buffer.getvalue()


# ============================================================
# 유틸리티 함수
# ============================================================

def format_timestamp(timestamp: int) -> str:
    """타임스탬프를 읽기 쉬운 형식으로 변환"""

    try:
        # milliseconds to seconds
        dt = datetime.fromtimestamp(timestamp / 1000)
        now = datetime.now()
        diff = now - dt

        if diff.days > 0:
            return f"{diff.days}일 전"
        elif diff.seconds >= 3600:
            return f"{diff.seconds // 3600}시간 전"
        elif diff.seconds >= 60:
            return f"{diff.seconds // 60}분 전"
        else:
            return "방금 전"
    except:
        return ""


def get_scene_range_options(total_scenes: int) -> List[tuple]:
    """씬 범위 선택 옵션 생성"""

    options = []

    # 50개씩 구간
    for start in range(1, total_scenes + 1, 50):
        end = min(start + 49, total_scenes)
        options.append((start, end, f"{start}-{end}"))

    return options


# ============================================================
# 한글 프롬프트 추천 씬 음영 표시 엑셀 (v1.1)
# ============================================================

def generate_prompts_excel_with_highlight(
    prompts: List[Dict[str, Any]],
    korean_recommended_scenes: set,
    highlight_color: str = "FFFF99",  # 연한 노란색
    include_recommendation_info: bool = True
) -> bytes:
    """한글 추천 씬 음영 표시된 엑셀 생성

    Args:
        prompts: 프롬프트 데이터 리스트
        korean_recommended_scenes: 한글 프롬프트 추천 씬 번호 집합
        highlight_color: 음영 색상 (HEX, 기본: 연한 노란색)
        include_recommendation_info: 추천 정보 컬럼 포함 여부

    Returns:
        엑셀 파일 bytes
    """

    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # 데이터 준비
    data = []
    for p in prompts:
        scene_num = p["scene_number"]
        is_korean_recommended = scene_num in korean_recommended_scenes

        row = {
            "씬번호": scene_num,
            "한글추천": "O" if is_korean_recommended else "",
            "나레이션": (p.get("narration", "")[:80] + "...") if len(p.get("narration", "")) > 80 else p.get("narration", ""),
            "원본프롬프트": p.get("original_prompt", ""),
            "한글프롬프트": p.get("korean_prompt", ""),
            "최종프롬프트": p.get("final_prompt", ""),
        }

        if include_recommendation_info:
            row["추천이유"] = p.get("recommendation_reason", "")
            row["추천문구"] = p.get("suggested_korean_text", "")

        data.append(row)

    df = pd.DataFrame(data)

    # 엑셀 생성
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='프롬프트')

        # 워크시트 가져오기
        workbook = writer.book
        worksheet = writer.sheets['프롬프트']

        # 스타일 정의
        highlight_fill = PatternFill(
            start_color=highlight_color,
            end_color=highlight_color,
            fill_type="solid"
        )

        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )

        header_font = Font(color="FFFFFF", bold=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 헤더 스타일 적용
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # 데이터 행 스타일 적용
        for row_num, row_data in enumerate(data, 2):  # 2부터 (헤더가 1)
            scene_num = row_data["씬번호"]
            is_recommended = scene_num in korean_recommended_scenes

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border

                # 한글 추천 씬이면 음영 적용
                if is_recommended:
                    cell.fill = highlight_fill

        # 컬럼 너비 조정
        column_widths = [
            8,   # 씬번호
            10,  # 한글추천
            40,  # 나레이션
            60,  # 원본프롬프트
            60,  # 한글프롬프트
            80,  # 최종프롬프트
        ]

        if include_recommendation_info:
            column_widths.extend([25, 30])  # 추천이유, 추천문구

        for col_idx, width in enumerate(column_widths, 1):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = width

        # 필터 추가
        worksheet.auto_filter.ref = worksheet.dimensions

        # 틀 고정 (헤더 행)
        worksheet.freeze_panes = 'A2'

    output.seek(0)
    return output.getvalue()


def collect_prompts_with_korean_recommendation(
    selected_scene_nums: List[int],
    scenes: List[dict],
    latest_images: Dict[int, dict],
    korean_recommended: Dict[str, Any],
    use_latest_metadata: bool = True,
    style_prefix: str = "",
    style_suffix: str = ""
) -> List[Dict[str, Any]]:
    """추천 정보가 포함된 프롬프트 수집

    Args:
        selected_scene_nums: 선택된 씬 번호 리스트
        scenes: 전체 씬 데이터
        latest_images: 씬별 최신 이미지 정보
        korean_recommended: AI 추천 결과 {"selected_scenes": [...], ...}
        use_latest_metadata: 최신 메타데이터 사용 여부
        style_prefix: 스타일 prefix
        style_suffix: 스타일 suffix

    Returns:
        추천 정보가 포함된 프롬프트 데이터 리스트
    """

    # 추천 씬 정보를 딕셔너리로 변환
    recommendation_map = {}
    for item in korean_recommended.get("selected_scenes", []):
        scene_num = item.get("scene_number")
        if scene_num:
            recommendation_map[scene_num] = {
                "reason": item.get("reason", ""),
                "suggested_text": item.get("suggested_korean_text", "")
            }

    prompts = []

    for scene_num in sorted(selected_scene_nums):
        # 씬 데이터 찾기
        scene_data = None

        for s in scenes:
            s_id = s.get("scene_id") or s.get("scene_num") or s.get("id")
            if s_id == scene_num:
                scene_data = s
                break

        if scene_data is None:
            scene_idx = scene_num - 1
            if 0 <= scene_idx < len(scenes):
                scene_data = scenes[scene_idx]

        if scene_data:
            prompt_data = get_prompt_for_scene(
                scene_num=scene_num,
                scene_data=scene_data,
                latest_images=latest_images,
                use_latest_metadata=use_latest_metadata,
                style_prefix=style_prefix,
                style_suffix=style_suffix
            )

            # 추천 정보 추가
            if scene_num in recommendation_map:
                prompt_data["is_korean_recommended"] = True
                prompt_data["recommendation_reason"] = recommendation_map[scene_num]["reason"]
                prompt_data["suggested_korean_text"] = recommendation_map[scene_num]["suggested_text"]
            else:
                prompt_data["is_korean_recommended"] = False
                prompt_data["recommendation_reason"] = ""
                prompt_data["suggested_korean_text"] = ""

            prompts.append(prompt_data)

    return prompts
